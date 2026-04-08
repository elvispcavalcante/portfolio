"""
Automação de Relatórios Mensais — Elvis P.C.
============================================
Este script lê automaticamente todas as planilhas de uma pasta,
consolida os dados por mês e categoria, e gera um relatório Excel
formatado e pronto para envio.

Uso:
    python gerar_relatorio.py

Requisitos:
    pip install pandas openpyxl

Estrutura esperada de cada planilha de entrada:
    Colunas: Data | Categoria | Descrição | Valor | Tipo (Receita/Despesa)
"""

import os
import sys
import glob
from datetime import datetime

# Garante saída UTF-8 no terminal Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

_MESES_PT = {
    "January": "Janeiro", "February": "Fevereiro", "March": "Março",
    "April": "Abril", "May": "Maio", "June": "Junho",
    "July": "Julho", "August": "Agosto", "September": "Setembro",
    "October": "Outubro", "November": "Novembro", "December": "Dezembro",
}

def _mes_pt(dt: datetime) -> str:
    """Retorna 'Mês/Ano' com o nome do mês em português."""
    return f"{_MESES_PT[dt.strftime('%B')]}/{dt.strftime('%Y')}"

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


# ── CONFIGURAÇÕES ────────────────────────────────────────────────────────────

_BASE = os.path.dirname(os.path.abspath(__file__))
PASTA_ENTRADA  = os.path.join(_BASE, "dados_mensais")       # Pasta com os .xlsx de entrada
PASTA_SAIDA    = os.path.join(_BASE, "relatorios_gerados")  # Pasta onde o relatório será salvo
NOME_RELATORIO = f"relatorio_{datetime.now().strftime('%Y%m')}_{_MESES_PT[datetime.now().strftime('%B')]}.xlsx"

# Cores
COR_AZUL_ESCURO  = "1F3864"
COR_AZUL_MEDIO   = "2E5597"
COR_AZUL_CLARO   = "BDD7EE"
COR_VERDE        = "375623"
COR_VERDE_CLARO  = "E2EFDA"
COR_VERMELHO     = "C00000"
COR_VERMELHO_CL  = "FFCCCC"
COR_AMARELO      = "FFF2CC"
COR_CINZA_CLARO  = "F2F2F2"
COR_BRANCO       = "FFFFFF"

borda = Border(
    left=Side(style='thin', color="BFBFBF"),
    right=Side(style='thin', color="BFBFBF"),
    top=Side(style='thin', color="BFBFBF"),
    bottom=Side(style='thin', color="BFBFBF"),
)


# ── FUNÇÕES AUXILIARES ───────────────────────────────────────────────────────

def aplicar_cabecalho(cell, texto, bg=COR_AZUL_MEDIO, cor_fonte=COR_BRANCO,
                      tamanho=11, negrito=True, centralizar=True):
    """Aplica estilo de cabeçalho a uma célula."""
    cell.value = texto
    cell.font = Font(name="Arial", bold=negrito, size=tamanho, color=cor_fonte)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(
        horizontal="center" if centralizar else "left",
        vertical="center", wrap_text=True
    )
    cell.border = borda


def aplicar_dado(cell, valor, negrito=False, fmt=None, cor_bg=COR_BRANCO,
                 cor_fonte="000000", centralizar=False):
    """Aplica estilo de dado comum a uma célula."""
    cell.value = valor
    cell.font = Font(name="Arial", bold=negrito, size=10, color=cor_fonte)
    cell.fill = PatternFill("solid", fgColor=cor_bg)
    cell.alignment = Alignment(
        horizontal="center" if centralizar else "left",
        vertical="center"
    )
    if fmt:
        cell.number_format = fmt
    cell.border = borda


# ── ETAPA 1: LEITURA E CONSOLIDAÇÃO DOS DADOS ────────────────────────────────

def ler_planilhas(pasta: str) -> pd.DataFrame:
    """
    Lê todos os arquivos .xlsx da pasta especificada e consolida em um
    único DataFrame.
    """
    arquivos = glob.glob(os.path.join(pasta, "*.xlsx"))

    if not arquivos:
        raise FileNotFoundError(
            f"Nenhum arquivo .xlsx encontrado em '{pasta}'.\n"
            "Execute o script com os arquivos de dados na pasta correta."
        )

    print(f"  → {len(arquivos)} arquivo(s) encontrado(s) em '{pasta}'")

    frames = []
    for arq in arquivos:
        print(f"     Lendo: {os.path.basename(arq)}")
        df = pd.read_excel(arq, parse_dates=["Data"])
        df["Arquivo Origem"] = os.path.basename(arq)
        frames.append(df)

    df_total = pd.concat(frames, ignore_index=True)
    print(f"  → Total de registros consolidados: {len(df_total)}\n")
    return df_total


def validar_colunas(df: pd.DataFrame) -> None:
    """Verifica se o DataFrame possui as colunas esperadas."""
    colunas_esperadas = {"Data", "Categoria", "Descrição", "Valor", "Tipo"}
    faltando = colunas_esperadas - set(df.columns)
    if faltando:
        raise ValueError(
            f"Colunas ausentes nos dados: {faltando}\n"
            f"Colunas encontradas: {list(df.columns)}"
        )


def limpar_dados(df: pd.DataFrame) -> pd.DataFrame:
    """
    Realiza limpeza básica: remove duplicatas, trata nulos,
    padroniza texto e extrai Mês/Ano.
    """
    df = df.drop_duplicates()
    df["Data"] = pd.to_datetime(df["Data"], errors="coerce")
    df = df.dropna(subset=["Data", "Valor", "Tipo"])
    df["Categoria"] = df["Categoria"].str.strip().str.title()
    df["Tipo"] = df["Tipo"].str.strip().str.capitalize()
    df["Valor"] = pd.to_numeric(df["Valor"], errors="coerce").fillna(0)
    df["Mês"]  = df["Data"].dt.month
    df["Ano"]  = df["Data"].dt.year
    df["Mês_Nome"] = (
        df["Data"].dt.strftime("%B").map(_MESES_PT) + "/" + df["Data"].dt.strftime("%Y")
    )
    return df


# ── ETAPA 2: ANÁLISE DOS DADOS ────────────────────────────────────────────────

def calcular_resumo(df: pd.DataFrame) -> dict:
    """Retorna um dicionário com as principais métricas do período."""
    receitas  = df[df["Tipo"] == "Receita"]["Valor"].sum()
    despesas  = df[df["Tipo"] == "Despesa"]["Valor"].sum()
    saldo     = receitas - despesas
    n_registros = len(df)

    por_categoria = (
        df.groupby(["Categoria", "Tipo"])["Valor"]
        .sum()
        .reset_index()
        .sort_values("Valor", ascending=False)
    )

    por_mes = (
        df.groupby(["Mês_Nome", "Mês", "Tipo"])["Valor"]
        .sum()
        .reset_index()
        .sort_values("Mês")
    )

    return {
        "total_receitas":   receitas,
        "total_despesas":   despesas,
        "saldo":            saldo,
        "n_registros":      n_registros,
        "por_categoria":    por_categoria,
        "por_mes":          por_mes,
        "df_completo":      df,
    }


# ── ETAPA 3: GERAÇÃO DO RELATÓRIO EXCEL ──────────────────────────────────────

def gerar_excel(resumo: dict, caminho_saida: str) -> None:
    """Gera o relatório Excel completo a partir do dicionário de resumo."""
    wb = Workbook()

    _criar_aba_resumo(wb, resumo)
    _criar_aba_por_categoria(wb, resumo["por_categoria"])
    _criar_aba_evolucao_mensal(wb, resumo["por_mes"])
    _criar_aba_dados_brutos(wb, resumo["df_completo"])

    os.makedirs(os.path.dirname(caminho_saida) or ".", exist_ok=True)
    wb.save(caminho_saida)
    print(f"  ✓ Relatório salvo em: {caminho_saida}")


def _criar_aba_resumo(wb: Workbook, resumo: dict) -> None:
    """Cria a aba 'Resumo Executivo' com KPIs e totais."""
    ws = wb.active
    ws.title = "Resumo Executivo"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = COR_AZUL_ESCURO

    ws.merge_cells("A1:F1")
    aplicar_cabecalho(ws["A1"],
                      f"RELATÓRIO FINANCEIRO CONSOLIDADO — {_mes_pt(datetime.now()).upper()}",
                      bg=COR_AZUL_ESCURO, tamanho=14)
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:F2")
    aplicar_cabecalho(ws["A2"],
                      f"Gerado automaticamente em {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
                      bg=COR_AZUL_MEDIO, tamanho=9, negrito=False)
    ws.row_dimensions[2].height = 18

    # KPIs
    kpis = [
        ("TOTAL DE RECEITAS",  resumo["total_receitas"],  COR_VERDE,    "R$ #,##0.00"),
        ("TOTAL DE DESPESAS",  resumo["total_despesas"],  COR_VERMELHO, "R$ #,##0.00"),
        ("SALDO DO PERÍODO",   resumo["saldo"],           COR_AZUL_MEDIO,"R$ #,##0.00"),
        ("REGISTROS ANALISADOS", resumo["n_registros"],  "595959",     "#,##0"),
    ]

    for i, (label, valor, cor, fmt) in enumerate(kpis):
        col = i + 1
        ws.merge_cells(f"{get_column_letter(col)}4:{get_column_letter(col)}4")
        ws.merge_cells(f"{get_column_letter(col)}5:{get_column_letter(col)}5")
        aplicar_cabecalho(ws.cell(4, col), label, bg=cor, tamanho=9, negrito=False)
        c = ws.cell(5, col)
        c.value = valor
        c.font = Font(name="Arial", bold=True, size=18, color=cor)
        c.fill = PatternFill("solid", fgColor=COR_CINZA_CLARO)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt
        c.border = borda
        ws.row_dimensions[4].height = 22
        ws.row_dimensions[5].height = 44

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 24


def _criar_aba_por_categoria(wb: Workbook, df: pd.DataFrame) -> None:
    """Cria a aba com análise por categoria."""
    ws = wb.create_sheet("Por Categoria")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = COR_AZUL_MEDIO

    ws.merge_cells("A1:D1")
    aplicar_cabecalho(ws["A1"], "TOTAL POR CATEGORIA", bg=COR_AZUL_ESCURO, tamanho=13)
    ws.row_dimensions[1].height = 32

    cabecalhos = ["Categoria", "Tipo", "Total (R$)", "% do Total"]
    for col, cab in enumerate(cabecalhos, 1):
        aplicar_cabecalho(ws.cell(2, col), cab)
    ws.row_dimensions[2].height = 22

    total_geral = df["Valor"].sum()
    for i, (_, row) in enumerate(df.iterrows()):
        r = i + 3
        bg = COR_CINZA_CLARO if i % 2 == 0 else COR_BRANCO
        cor_tipo = COR_VERDE_CLARO if row["Tipo"] == "Receita" else COR_VERMELHO_CL

        aplicar_dado(ws.cell(r, 1), row["Categoria"], negrito=True, cor_bg=bg)
        aplicar_dado(ws.cell(r, 2), row["Tipo"], centralizar=True, cor_bg=cor_tipo)
        aplicar_dado(ws.cell(r, 3), row["Valor"], fmt="R$ #,##0.00",
                     centralizar=True, cor_bg=bg)
        pct = row["Valor"] / total_geral if total_geral else 0
        aplicar_dado(ws.cell(r, 4), pct, fmt="0.0%", centralizar=True, cor_bg=bg)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 16


def _criar_aba_evolucao_mensal(wb: Workbook, df: pd.DataFrame) -> None:
    """Cria a aba de evolução mensal com gráfico."""
    ws = wb.create_sheet("Evolução Mensal")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "375623"

    ws.merge_cells("A1:D1")
    aplicar_cabecalho(ws["A1"], "EVOLUÇÃO MENSAL — RECEITAS E DESPESAS",
                      bg=COR_AZUL_ESCURO, tamanho=13)
    ws.row_dimensions[1].height = 32

    for col, cab in enumerate(["Mês", "Tipo", "Total (R$)"], 1):
        aplicar_cabecalho(ws.cell(2, col), cab)

    for i, (_, row) in enumerate(df.iterrows()):
        r = i + 3
        bg = COR_CINZA_CLARO if i % 2 == 0 else COR_BRANCO
        aplicar_dado(ws.cell(r, 1), row["Mês_Nome"], cor_bg=bg)
        cor_tipo = COR_VERDE_CLARO if row["Tipo"] == "Receita" else COR_VERMELHO_CL
        aplicar_dado(ws.cell(r, 2), row["Tipo"], centralizar=True, cor_bg=cor_tipo)
        aplicar_dado(ws.cell(r, 3), row["Valor"], fmt="R$ #,##0.00",
                     centralizar=True, cor_bg=bg)

    # Gráfico
    n_linhas = len(df) + 2
    chart = BarChart()
    chart.type = "col"
    chart.title = "Receitas × Despesas por Mês"
    chart.y_axis.title = "Valor (R$)"
    chart.style = 10
    chart.width = 22
    chart.height = 14
    data  = Reference(ws, min_col=3, min_row=2, max_row=n_linhas)
    cats  = Reference(ws, min_col=1, min_row=3, max_row=n_linhas)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "E2")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20


def _criar_aba_dados_brutos(wb: Workbook, df: pd.DataFrame) -> None:
    """Cria a aba com todos os registros consolidados."""
    ws = wb.create_sheet("Dados Consolidados")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "595959"

    ws.merge_cells("A1:G1")
    aplicar_cabecalho(ws["A1"], "BASE DE DADOS CONSOLIDADA",
                      bg=COR_AZUL_ESCURO, tamanho=13)
    ws.row_dimensions[1].height = 32

    colunas_exib = ["Data", "Categoria", "Descrição", "Valor", "Tipo", "Arquivo Origem"]
    for col, cab in enumerate(colunas_exib, 1):
        aplicar_cabecalho(ws.cell(2, col), cab)

    df_exib = df[colunas_exib].copy()
    for i, (_, row) in enumerate(df_exib.iterrows()):
        r = i + 3
        bg = COR_CINZA_CLARO if i % 2 == 0 else COR_BRANCO
        aplicar_dado(ws.cell(r, 1), row["Data"].strftime("%d/%m/%Y"),
                     centralizar=True, cor_bg=bg)
        aplicar_dado(ws.cell(r, 2), row["Categoria"], negrito=True, cor_bg=bg)
        aplicar_dado(ws.cell(r, 3), row["Descrição"], cor_bg=bg)
        aplicar_dado(ws.cell(r, 4), row["Valor"], fmt="R$ #,##0.00",
                     centralizar=True, cor_bg=bg)
        cor_tipo = COR_VERDE_CLARO if row["Tipo"] == "Receita" else COR_VERMELHO_CL
        aplicar_dado(ws.cell(r, 5), row["Tipo"], centralizar=True, cor_bg=cor_tipo)
        aplicar_dado(ws.cell(r, 6), row["Arquivo Origem"], cor_bg=bg)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 28


# ── ETAPA 4: EXECUÇÃO PRINCIPAL ───────────────────────────────────────────────

def main():
    print("=" * 58)
    print("  GERADOR DE RELATÓRIOS — Elvis P.C.")
    print("=" * 58)

    print("\n[1/4] Lendo planilhas de entrada...")
    df_raw = ler_planilhas(PASTA_ENTRADA)

    print("[2/4] Validando e limpando os dados...")
    validar_colunas(df_raw)
    df_limpo = limpar_dados(df_raw)
    print(f"  → {len(df_limpo)} registros válidos após limpeza\n")

    print("[3/4] Calculando resumo e métricas...")
    resumo = calcular_resumo(df_limpo)
    print(f"  → Receitas:  R$ {resumo['total_receitas']:,.2f}")
    print(f"  → Despesas:  R$ {resumo['total_despesas']:,.2f}")
    print(f"  → Saldo:     R$ {resumo['saldo']:,.2f}\n")

    print("[4/4] Gerando relatório Excel...")
    caminho_saida = os.path.join(PASTA_SAIDA, NOME_RELATORIO)
    gerar_excel(resumo, caminho_saida)

    print("\n" + "=" * 58)
    print("  Relatório gerado com sucesso!")
    print(f"  Arquivo: {caminho_saida}")
    print("=" * 58)


if __name__ == "__main__":
    main()
